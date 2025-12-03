import os
import re
import pdfplumber
import pytesseract
from pdf2image import convert_from_path
import openpyxl
import unicodedata
import pandas as pd

# >>> Ajuste aqui para o caminho da pasta bin do Poppler <<<
POPPLER_PATH = r"C:\Users\rafae\Downloads\Release-25.11.0-0\poppler-25.11.0\Library\bin"
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR"


# Regex para CPF (aceita diferentes formatos - mais flexível)
# Aceita: 123.456.789-10, 12345678910, 123 456 789 10, 123.456.789/10, etc.
padrao_cpf = re.compile(r'(\d{3}[\.\s\/-]?\d{3}[\.\s\/-]?\d{3}[\.\s\/-]?\d{2})')

# Regex para CNPJ (aceita diferentes formatos)
# Aceita: 12.345.678/0001-90, 12345678000190, 12 345 678 0001 90, etc.
padrao_cnpj = re.compile(r'(\d{2}[\.\s\/-]?\d{3}[\.\s\/-]?\d{3}[\.\s\/-]?\d{4}[\.\s\/-]?\d{2})')

# Regex para encontrar campo CPF/CNPJ (aceita diferentes variações: CPF, cpf, C.P.F, c.p.f, CNPJ, etc.)
padrao_campo_cpf = re.compile(r'c\.?\s*p\.?\s*f\.?|c\.?\s*n\.?\s*p\.?\s*j\.?', re.IGNORECASE)

def normalizar_nome(nome):
    """Normaliza o nome para comparação (remove espaços extras, converte para minúsculas, remove acentos)"""
    if not nome:
        return ""
    # Remove acentos
    nome_sem_acentos = ''.join(
        c for c in unicodedata.normalize('NFD', str(nome))
        if unicodedata.category(c) != 'Mn'
    )
    # Remove espaços extras e converte para minúsculas
    return re.sub(r'\s+', ' ', nome_sem_acentos.strip().lower())

def encontrar_cpf_proximo_ao_nome(texto, nome_requerente, debug=False):
    """Procura pelo nome do requerente no texto e retorna o CPF/CNPJ mais próximo"""
    # Dividir texto em linhas
    linhas = texto.split('\n')
    
    # Primeiro, verificar se há CPFs e CNPJs no texto
    todos_cpfs = padrao_cpf.findall(texto)
    todos_cnpjs = padrao_cnpj.findall(texto)
    todos_documentos = todos_cpfs + todos_cnpjs  # Combinar CPFs e CNPJs
    
    if debug:
        print(f"    [DEBUG] Total de CPFs encontrados no texto: {len(todos_cpfs)}")
        if todos_cpfs:
            print(f"    [DEBUG] CPFs encontrados: {todos_cpfs}")
        print(f"    [DEBUG] Total de CNPJs encontrados no texto: {len(todos_cnpjs)}")
        if todos_cnpjs:
            print(f"    [DEBUG] CNPJs encontrados: {todos_cnpjs}")
    
    # ESTRATÉGIA ESPECIAL: Procurar padrão "Nome:" seguido de "CPF/CNPJ:" (comum em Anexo II)
    # Esta estratégia funciona mesmo sem o nome do requerente
    for i, linha in enumerate(linhas):
        linha_lower = linha.lower()
        # Procurar linha com "Nome:" ou "nome:"
        if "nome:" in linha_lower:
            # Procurar CPF/CNPJ nas próximas 3 linhas (geralmente está logo após)
            for j in range(i, min(i+4, len(linhas))):
                linha_doc = linhas[j]
                # Verificar se tem "CPF" ou "CPF/CNPJ" na linha
                if padrao_campo_cpf.search(linha_doc):
                    # Procurar CPF primeiro
                    cpfs = padrao_cpf.findall(linha_doc)
                    if cpfs:
                        if debug:
                            print(f"    [DEBUG] CPF encontrado apos campo 'Nome:' na linha {j+1}: {cpfs[0]}")
                        return cpfs[0]
                    # Se não encontrou CPF, procurar CNPJ
                    cnpjs = padrao_cnpj.findall(linha_doc)
                    if cnpjs:
                        if debug:
                            print(f"    [DEBUG] CNPJ encontrado apos campo 'Nome:' na linha {j+1}: {cnpjs[0]}")
                        return cnpjs[0]
                # Ou procurar CPF/CNPJ diretamente nas linhas seguintes
                cpfs = padrao_cpf.findall(linha_doc)
                if cpfs:
                    if debug:
                        print(f"    [DEBUG] CPF encontrado apos 'Nome:' na linha {j+1}: {cpfs[0]}")
                    return cpfs[0]
                cnpjs = padrao_cnpj.findall(linha_doc)
                if cnpjs:
                    if debug:
                        print(f"    [DEBUG] CNPJ encontrado apos 'Nome:' na linha {j+1}: {cnpjs[0]}")
                    return cnpjs[0]
    
    # Se não há nome do requerente fornecido, retornar primeiro CPF/CNPJ encontrado
    nome_normalizado = normalizar_nome(nome_requerente) if nome_requerente else ""
    if not nome_normalizado:
        if debug:
            print("    [DEBUG] Nome do requerente vazio - retornando primeiro CPF/CNPJ encontrado")
        # Retornar CPF primeiro, se não houver, retornar CNPJ
        return todos_cpfs[0] if todos_cpfs else (todos_cnpjs[0] if todos_cnpjs else None)
    
    # Procurar linha(s) com o nome do requerente
    indices_nome = []
    palavras_nome = nome_normalizado.split()
    
    # Busca mais flexível: aceita nomes com 1 palavra também, mas prioriza nomes com mais palavras
    for i, linha in enumerate(linhas):
        linha_normalizada = normalizar_nome(linha)
        # Verifica quantas palavras do nome estão na linha
        palavras_encontradas = sum(1 for palavra in palavras_nome if palavra in linha_normalizada)
        
        # Se encontrou pelo menos 1 palavra (para nomes curtos) ou 2+ palavras (para nomes longos)
        if len(palavras_nome) == 1:
            if palavras_encontradas >= 1:
                indices_nome.append(i)
        else:
            if palavras_encontradas >= 2:
                indices_nome.append(i)
    
    if debug:
        print(f"    [DEBUG] Nome normalizado: '{nome_normalizado}'")
        print(f"    [DEBUG] Linhas com nome encontradas: {indices_nome}")
        if indices_nome:
            print(f"    [DEBUG] Exemplo de linha com nome: '{linhas[indices_nome[0]]}'")
    
    # Se não encontrou o nome, mas há CPFs/CNPJs, retornar o primeiro (fallback)
    if not indices_nome:
        if debug:
            print("    [DEBUG] Nome não encontrado no texto, mas tentando retornar primeiro CPF/CNPJ")
        return todos_cpfs[0] if todos_cpfs else (todos_cnpjs[0] if todos_cnpjs else None)
    
    # Usar o primeiro índice encontrado como referência
    indice_nome = indices_nome[0]
    
    # Procurar CPF nas linhas próximas ao nome (até 15 linhas antes e depois)
    inicio_busca = max(0, indice_nome - 15)
    fim_busca = min(len(linhas), indice_nome + 16)
    
    # Estratégia 1: Procurar por padrão "CPF/CNPJ" seguido de números na região próxima
    for i in range(inicio_busca, fim_busca):
        linha = linhas[i]
        match_campo = padrao_campo_cpf.search(linha)
        if match_campo:
            # Se encontrou o campo CPF/CNPJ, procurar o número na mesma linha ou próximas
            texto_proximo = ' '.join(linhas[max(0, i-2):min(len(linhas), i+4)])
            cpfs = padrao_cpf.findall(texto_proximo)
            if cpfs:
                if debug:
                    print(f"    [DEBUG] CPF encontrado próximo ao campo 'CPF/CNPJ' na linha {i}: {cpfs[0]}")
                return cpfs[0]
            cnpjs = padrao_cnpj.findall(texto_proximo)
            if cnpjs:
                if debug:
                    print(f"    [DEBUG] CNPJ encontrado próximo ao campo 'CPF/CNPJ' na linha {i}: {cnpjs[0]}")
                return cnpjs[0]
    
    # Estratégia 2: Procurar CPF/CNPJ após o nome (nas linhas seguintes)
    for i in range(indice_nome, fim_busca):
        linha = linhas[i]
        cpfs = padrao_cpf.findall(linha)
        if cpfs:
            if debug:
                print(f"    [DEBUG] CPF encontrado na linha {i} (após o nome): {cpfs[0]}")
            return cpfs[0]
        cnpjs = padrao_cnpj.findall(linha)
        if cnpjs:
            if debug:
                print(f"    [DEBUG] CNPJ encontrado na linha {i} (após o nome): {cnpjs[0]}")
            return cnpjs[0]
    
    # Estratégia 3: Procurar CPF/CNPJ em um bloco de texto próximo ao nome
    texto_proximo = ' '.join(linhas[inicio_busca:fim_busca])
    cpfs = padrao_cpf.findall(texto_proximo)
    if cpfs:
        if debug:
            print(f"    [DEBUG] CPF encontrado no bloco próximo ao nome: {cpfs[0]}")
        return cpfs[0]
    cnpjs = padrao_cnpj.findall(texto_proximo)
    if cnpjs:
        if debug:
            print(f"    [DEBUG] CNPJ encontrado no bloco próximo ao nome: {cnpjs[0]}")
        return cnpjs[0]
    
    # Fallback: procurar qualquer CPF/CNPJ na página (último recurso)
    if todos_cpfs:
        if debug:
            print(f"    [DEBUG] Retornando primeiro CPF encontrado (fallback): {todos_cpfs[0]}")
        return todos_cpfs[0]
    if todos_cnpjs:
        if debug:
            print(f"    [DEBUG] Retornando primeiro CNPJ encontrado (fallback): {todos_cnpjs[0]}")
        return todos_cnpjs[0]
    
    if debug:
        print("    [DEBUG] Nenhum CPF encontrado")
    return None

def extrair_todas_partes_anexoII(caminho_pdf, debug=False):
    """Extrai todas as partes (Requerente, Invitante, Interessado, etc.) e seus CPFs/CNPJs do Anexo II.
    Retorna um dicionário com {tipo_parte: {'nome': str, 'cpf_cnpj': str}}"""
    pagina_anexo = None
    texto_pagina = ""
    
    # Tipos de partes a procurar
    tipos_partes = [
        'requerente', 'invitante', 'interessado', 'cedente', 'sucessora', 
        'favorecido', 'sucessor', 'cessionário', 'favorecida'
    ]
    
    # 1. Procurar página do Anexo II
    with pdfplumber.open(caminho_pdf) as pdf:
        for i, pagina in enumerate(pdf.pages):
            texto = pagina.extract_text()
            if texto and ("Anexo II" in texto or "ANEXO II" in texto or "anexo ii" in texto):
                pagina_anexo = i
                texto_pagina = texto
                # Tentar também extrair tabelas
                tabelas = pagina.extract_tables()
                if tabelas:
                    for tabela in tabelas:
                        for linha in tabela:
                            if linha:
                                texto_pagina += "\n" + " ".join([str(cell) if cell else "" for cell in linha])
                break
    
    if pagina_anexo is None:
        if debug:
            print("    [DEBUG] Página 'Anexo II' não encontrada no PDF")
        return {}
    
    if debug:
        print(f"    [DEBUG] Página Anexo II encontrada: página {pagina_anexo + 1}")
        print(f"    [DEBUG] Tamanho do texto extraído: {len(texto_pagina)} caracteres")
    
    # Se não achou texto, tentar OCR
    if len(texto_pagina) < 100:
        if debug:
            print("    [DEBUG] Texto muito curto, tentando OCR...")
        try:
            paginas = convert_from_path(
                caminho_pdf,
                first_page=pagina_anexo+1,
                last_page=pagina_anexo+1,
                poppler_path=POPPLER_PATH,
                dpi=300
            )
            texto_pagina = pytesseract.image_to_string(paginas[0], lang="por")
        except Exception as e:
            if debug:
                print(f"    [DEBUG] Erro no OCR: {e}")
    
    # Extrair todas as partes
    partes_encontradas = {}
    linhas = texto_pagina.split('\n')
    
    for tipo in tipos_partes:
        # Procurar por padrões como "Requerente:", "Invitante:", etc.
        # Incluir variações de gênero (ex: Interessado/Interessada)
        padroes = [
            f"{tipo}:",
            f"{tipo.capitalize()}:",
            f"{tipo.upper()}:",
            f"{tipo} n°",
            f"{tipo.capitalize()} n°",
            f"{tipo} n.",
            f"{tipo.capitalize()} n.",
        ]
        
        # Adicionar variações de gênero para alguns tipos
        if tipo == 'interessado':
            padroes.extend(['interessada:', 'Interessada:', 'INTERESSADA:'])
        elif tipo == 'favorecido':
            padroes.extend(['favorecida:', 'Favorecida:', 'FAVORECIDA:'])
        
        for i, linha in enumerate(linhas):
            linha_lower = linha.lower()
            
            # Verificar se encontrou algum padrão
            encontrou_tipo = False
            for padrao in padroes:
                if padrao.lower() in linha_lower:
                    encontrou_tipo = True
                    break
            
            if encontrou_tipo:
                # Extrair nome (pode estar na mesma linha ou nas próximas)
                nome = ""
                cpf_cnpj = None
                
                # Tentar extrair nome da mesma linha
                if ":" in linha:
                    partes_linha = linha.split(":", 1)
                    if len(partes_linha) > 1:
                        nome_candidato = partes_linha[1].strip()
                        # Remover números no início (ex: "1. Nome da Pessoa")
                        nome_candidato = re.sub(r'^\d+[\.\)]\s*', '', nome_candidato)
                        # Se o nome candidato não contém apenas números ou campos conhecidos
                        if nome_candidato and len(nome_candidato) > 3 and not re.match(r'^\d+$', nome_candidato):
                            if not any(campo in nome_candidato.lower() for campo in ['cpf', 'cnpj', 'data', 'valor']):
                                nome = nome_candidato
                
                # Se não encontrou nome, procurar nas próximas linhas
                if not nome or len(nome) < 3:
                    for j in range(i+1, min(i+6, len(linhas))):
                        linha_seguinte = linhas[j].strip()
                        # Remover números no início
                        linha_seguinte = re.sub(r'^\d+[\.\)]\s*', '', linha_seguinte)
                        # Se a linha seguinte não começa com um campo conhecido, é provavelmente o nome
                        if linha_seguinte and len(linha_seguinte) > 3:
                            if not any(campo in linha_seguinte.lower() for campo in 
                                ['cpf', 'cnpj', 'data', 'valor', 'contribuições', 'anexo', 'nascimento']):
                                if not re.match(r'^\d+[\.\s\/-]*$', linha_seguinte):  # Não é só números
                                    nome = linha_seguinte
                                    break
                
                # Procurar CPF/CNPJ nas próximas 5 linhas
                for j in range(i, min(i+6, len(linhas))):
                    linha_doc = linhas[j]
                    
                    # Verificar se tem campo CPF/CNPJ
                    if padrao_campo_cpf.search(linha_doc):
                        cpfs = padrao_cpf.findall(linha_doc)
                        if cpfs:
                            cpf_cnpj = cpfs[0]
                            break
                        cnpjs = padrao_cnpj.findall(linha_doc)
                        if cnpjs:
                            cpf_cnpj = cnpjs[0]
                            break
                    else:
                        # Procurar CPF/CNPJ diretamente
                        cpfs = padrao_cpf.findall(linha_doc)
                        if cpfs:
                            cpf_cnpj = cpfs[0]
                            break
                        cnpjs = padrao_cnpj.findall(linha_doc)
                        if cnpjs:
                            cpf_cnpj = cnpjs[0]
                            break
                
                # Se encontrou nome ou CPF/CNPJ, adicionar ao dicionário
                if nome or cpf_cnpj:
                    partes_encontradas[tipo] = {
                        'nome': nome if nome else '',
                        'cpf_cnpj': cpf_cnpj if cpf_cnpj else ''
                    }
                    if debug:
                        print(f"    [DEBUG] {tipo.capitalize()} encontrado: Nome='{nome}', CPF/CNPJ='{cpf_cnpj}'")
                    break  # Encontrou essa parte, passar para a próxima
    
    return partes_encontradas

def extrair_cpf_anexoII(caminho_pdf, nome_requerente, debug=False):
    """Extrai CPF ou CNPJ do requerente na página do Anexo II do PDF.
    Funciona mesmo se o nome do requerente estiver vazio."""
    pagina_anexo = None
    texto_pagina = ""

    # 1. Procurar página do Anexo II
    with pdfplumber.open(caminho_pdf) as pdf:
        for i, pagina in enumerate(pdf.pages):
            texto = pagina.extract_text()
            if texto and ("Anexo II" in texto or "ANEXO II" in texto or "anexo ii" in texto):
                pagina_anexo = i
                texto_pagina = texto
                # Tentar também extrair tabelas
                tabelas = pagina.extract_tables()
                if tabelas:
                    for tabela in tabelas:
                        for linha in tabela:
                            if linha:
                                texto_pagina += "\n" + " ".join([str(cell) if cell else "" for cell in linha])
                break

    if pagina_anexo is None:
        if debug:
            print("    [DEBUG] Página 'Anexo II' não encontrada no PDF")
        return None

    if debug:
        print(f"    [DEBUG] Página Anexo II encontrada: página {pagina_anexo + 1}")
        print(f"    [DEBUG] Tamanho do texto extraído: {len(texto_pagina)} caracteres")

    # 2. Tentar extrair CPF associado ao nome do requerente do texto extraído
    cpf = encontrar_cpf_proximo_ao_nome(texto_pagina, nome_requerente, debug=debug)
    if cpf:
        return cpf

    # 3. Se não achou, converter só a página do Anexo II em imagem e usar OCR
    if debug:
        print("    [DEBUG] Tentando OCR na página do Anexo II...")
    try:
        paginas = convert_from_path(
            caminho_pdf,
            first_page=pagina_anexo+1,
            last_page=pagina_anexo+1,
            poppler_path=POPPLER_PATH,
            dpi=300  # Maior resolução para melhor OCR
        )
        texto_ocr = pytesseract.image_to_string(paginas[0], lang="por")
        
        if debug:
            print(f"    [DEBUG] Tamanho do texto OCR: {len(texto_ocr)} caracteres")
        
        # Procurar CPF associado ao nome no texto do OCR
        cpf = encontrar_cpf_proximo_ao_nome(texto_ocr, nome_requerente, debug=debug)
        if cpf:
            return cpf
        
        # Fallback: retornar primeiro CPF/CNPJ encontrado se não conseguir associar ao nome
        cpfs = padrao_cpf.findall(texto_ocr)
        cnpjs = padrao_cnpj.findall(texto_ocr)
        if debug:
            if cpfs:
                print(f"    [DEBUG] CPFs encontrados no OCR (fallback): {cpfs}")
            if cnpjs:
                print(f"    [DEBUG] CNPJs encontrados no OCR (fallback): {cnpjs}")
        return cpfs[0] if cpfs else (cnpjs[0] if cnpjs else None)
    except Exception as e:
        if debug:
            print(f"    [DEBUG] Erro no OCR: {e}")
        return None


def atualizar_excel_iterativo(caminho_excel, pasta_pdfs):
    # Tentar abrir o arquivo Excel
    try:
        wb = openpyxl.load_workbook(caminho_excel)
    except PermissionError:
        print(f"\n[ERRO] O arquivo Excel está aberto ou sendo usado por outro programa!")
        print(f"Por favor, feche o arquivo '{caminho_excel}' e tente novamente.")
        return
    except FileNotFoundError:
        print(f"\n[ERRO] Arquivo Excel não encontrado: '{caminho_excel}'")
        return
    except Exception as e:
        print(f"\n[ERRO] Erro ao abrir o arquivo Excel: {e}")
        return
    
    ws = wb.active

    # Encontrar colunas existentes
    cabecalhos = [cell.value for cell in ws[1]]
    
    # Encontrar coluna do número do processo (geralmente primeira coluna)
    col_numero_processo = 1
    
    # Tipos de partes a procurar no Excel
    tipos_partes = [
        'requerente', 'invitante', 'interessado', 'cedente', 'sucessora', 
        'favorecido', 'sucessor', 'cessionário', 'favorecida'
    ]
    
    # Mapear colunas de partes e criar colunas de CPF/CNPJ
    colunas_partes = {}  # {tipo: {'nome': col_idx, 'cpf_cnpj': col_idx}}
    
    for tipo in tipos_partes:
        # Procurar coluna com o nome da parte
        col_nome = None
        for idx, cabecalho in enumerate(cabecalhos, start=1):
            if cabecalho:
                cabecalho_lower = str(cabecalho).lower()
                # Verificar se o cabeçalho contém o tipo de parte e não é coluna de CPF/CNPJ
                if tipo in cabecalho_lower and "cpf" not in cabecalho_lower and "cnpj" not in cabecalho_lower:
                    col_nome = idx
                    break
        
        if col_nome:
            colunas_partes[tipo] = {'nome': col_nome, 'cpf_cnpj': None}
            
            # Criar ou encontrar coluna de CPF/CNPJ ao lado da coluna de nome
            nome_col_cpf = f"{tipo.capitalize()} CPF/CNPJ"
            col_cpf = None
            
            # Verificar se a coluna já existe
            for idx, cabecalho in enumerate(cabecalhos, start=1):
                if cabecalho and (tipo in str(cabecalho).lower() and ("cpf" in str(cabecalho).lower() or "cnpj" in str(cabecalho).lower())):
                    col_cpf = idx
                    break
            
            if not col_cpf:
                # Inserir coluna após a coluna de nome
                ws.insert_cols(col_nome + 1)
                ws.cell(row=1, column=col_nome + 1, value=nome_col_cpf)
                col_cpf = col_nome + 1
                # Atualizar lista de cabeçalhos
                cabecalhos = [cell.value for cell in ws[1]]
            
            colunas_partes[tipo]['cpf_cnpj'] = col_cpf
    
    if not colunas_partes:
        print("[AVISO] Nenhuma coluna de parte encontrada no Excel!")
        print("Procurando por: Requerente, Invitante, Interessado, Cedente, Sucessora, Favorecido, Sucessor, Cessionário, Favorecida")
        return

    # Iterar linhas
    for row in range(2, ws.max_row+1):
        numero_processo = str(ws.cell(row=row, column=col_numero_processo).value).strip()
        caminho_pdf = os.path.join(pasta_pdfs, f"{numero_processo}.pdf")

        if os.path.exists(caminho_pdf):
            print(f"\n[...] Processo {numero_processo}: Processando...")
            # Ativar debug apenas para os primeiros processos
            debug = (row <= 3)
            
            # Extrair todas as partes do PDF
            partes_encontradas = extrair_todas_partes_anexoII(caminho_pdf, debug=debug)
            
            # Preencher as colunas correspondentes
            for tipo, dados in partes_encontradas.items():
                if tipo in colunas_partes:
                    col_cpf = colunas_partes[tipo]['cpf_cnpj']
                    if col_cpf and dados['cpf_cnpj']:
                        ws.cell(row=row, column=col_cpf, value=dados['cpf_cnpj'])
                        print(f"[OK] {tipo.capitalize()}: CPF/CNPJ encontrado {dados['cpf_cnpj']}")
                    elif col_cpf:
                        ws.cell(row=row, column=col_cpf, value="CPF/CNPJ não encontrado")
                        print(f"[!] {tipo.capitalize()}: CPF/CNPJ não encontrado")
            
            # Se não encontrou nenhuma parte, tentar método antigo (compatibilidade)
            if not partes_encontradas:
                print(f"[!] Nenhuma parte encontrada, tentando método alternativo...")
                # Tentar encontrar pelo nome do requerente se existir
                if 'requerente' in colunas_partes:
                    col_requerente = colunas_partes['requerente']['nome']
                    nome_requerente = str(ws.cell(row=row, column=col_requerente).value).strip() if ws.cell(row=row, column=col_requerente).value else ""
                    cpf = extrair_cpf_anexoII(caminho_pdf, nome_requerente, debug=debug)
                    if cpf and 'requerente' in colunas_partes:
                        col_cpf = colunas_partes['requerente']['cpf_cnpj']
                        if col_cpf:
                            ws.cell(row=row, column=col_cpf, value=cpf)
                            print(f"[OK] Requerente: CPF/CNPJ encontrado {cpf}")
        else:
            # Preencher todas as colunas de CPF/CNPJ com "PDF não encontrado"
            for tipo, dados in colunas_partes.items():
                if dados['cpf_cnpj']:
                    ws.cell(row=row, column=dados['cpf_cnpj'], value="PDF não encontrado")
            print(f"[X] Processo {numero_processo}: PDF não encontrado")

        # Salvar imediatamente após cada linha
        try:
            wb.save(caminho_excel)
        except PermissionError:
            print(f"\n[ERRO] Não foi possível salvar o arquivo Excel!")
            print(f"O arquivo '{caminho_excel}' pode estar aberto. Feche-o e execute o script novamente.")
            wb.close()
            return
        except Exception as e:
            print(f"\n[ERRO] Erro ao salvar o arquivo Excel: {e}")
            wb.close()
            return
    
    # Fechar o arquivo ao final
    wb.close()
    print("\n[CONCLUÍDO] Processamento finalizado com sucesso!")


# Exemplo de uso
pasta_pdfs = r"G:\Drives compartilhados\Tecnologia\PDFs - Esaj TJSP"
caminho_excel = r"C:\Users\rafae\OneDrive\Desktop\robo\resultados_processo_final.xlsx"

atualizar_excel_iterativo(caminho_excel, pasta_pdfs)
