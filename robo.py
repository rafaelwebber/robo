import os
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
import pandas as pd
import logging
from logging.handlers import RotatingFileHandler
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    NoSuchElementException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager


CAMINHO_PLANILHA = Path("XXXXXXX")
PASTA_DOWNLOAD = Path("XXXXXXX")
URL_LOGIN = (
    "https://esaj.tjsp.jus.br/sajcas/login"
    "?service=https%3A%2F%2Fesaj.tjsp.jus.br%2Fcpopg%2FabrirConsultaDeRequisitorios.do"
)
URL_CONSULTA = "https://esaj.tjsp.jus.br/cpopg/abrirConsultaDeRequisitorios.do"
TEMPO_DOWNLOAD = 90
#GAP_ENTRE_PROCESSOS = 3  # segundos: pequena pausa entre cada processo
#PAUSA_2_MINUTOS = 120  # segundos: pausa longa de 2 minutos entre processos
LOG_ARQUIVO = Path("erros_processos.log")
NOME_ARQUIVO_RESULTADOS = "resultados_processos.xlsx"
logger = logging.getLogger("robo_processos")

def carregar_processos(caminho: Path, limite: Optional[int] = None) -> List[str]:
    df = pd.read_excel(caminho)
    serie = df["Processo"].dropna().astype(str)
    if limite:
        serie = serie.iloc[:limite]
    return serie.tolist()


def separar_numero_processo(numero: str) -> Tuple[str, str]:
    numero_limpo = re.sub(r"[.-]", "", numero)
    if len(numero_limpo) != 20:
        raise ValueError(f"Número de processo inválido: {numero}")
    return numero_limpo[:13], numero_limpo[16:]


def inicializar_driver() -> Chrome:
    PASTA_DOWNLOAD.mkdir(parents=True, exist_ok=True)

    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-infobars")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    prefs = {
        "download.default_directory": str(PASTA_DOWNLOAD),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "plugins.always_open_pdf_externally": True,
    }
    options.add_experimental_option("prefs", prefs)

    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def aguardar(
    driver: Chrome,
    locator: Tuple[str, str],
    condicao=EC.presence_of_element_located,
    timeout: int = 20,
):
    return WebDriverWait(driver, timeout).until(condicao(locator))


def preencher_campo(driver: Chrome, locator: Tuple[str, str], valor: str, timeout: int = 15):
    campo = aguardar(driver, locator, EC.element_to_be_clickable, timeout)
    campo.click()
    campo.send_keys(Keys.CONTROL + "a")
    campo.send_keys(Keys.DELETE)
    campo.send_keys(valor)


def clicar_com_retentativa(
    driver: Chrome,
    locator: Tuple[str, str],
    timeout: int = 15,
    tentativas: int = 3,
):
    ultimo_erro: Optional[Exception] = None
    for tentativa in range(1, tentativas + 1):
        try:
            elemento = aguardar(driver, locator, EC.element_to_be_clickable, timeout)
            elemento.click()
            return
        except (ElementClickInterceptedException, StaleElementReferenceException, TimeoutException) as erro:
            ultimo_erro = erro
            time.sleep(tentativa)
    if ultimo_erro:
        raise ultimo_erro


def extrair_texto_por_id(driver, element_id, timeout = 10) -> str:
    try:
        # Aguarda o elemento estar presente no DOM
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, element_id)))
        prazo = time.time() + timeout

        while time.time() < prazo:
            elemento = driver.find_element(By.ID, element_id)
            texto = (elemento.text or "").strip()
            if texto:
                return texto
            texto = (elemento.get_attribute("textContent") or "").strip()
            if texto:
                return texto
            texto = (elemento.get_attribute("innerText") or "").strip()
            if texto:
                return texto
            time.sleep(0.3)

        return ""
    except TimeoutException:
        return ""

def extrair_html(driver: Chrome, locator: Tuple[str, str], timeout: int = 10) -> str:
    try:
        elemento = aguardar(driver, locator, EC.presence_of_element_located, timeout)
        return elemento.get_attribute("outerHTML") or ""
    except TimeoutException:
        return ""


def aguardar_resultado_consulta(driver: Chrome, timeout: int = 25):
    def _resultado_disponivel(_driver: Chrome):
        possui_classe = _driver.find_elements(By.ID, "classeProcesso")
        erro = _driver.find_elements(By.CSS_SELECTOR, ".mensagemErro")
        return bool(possui_classe or erro)

    WebDriverWait(driver, timeout).until(_resultado_disponivel)


def extrair_outros_numeros(html: str) -> str:
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    span_rotulo = soup.find("span", string=re.compile("Outros números", re.IGNORECASE))
    if not span_rotulo:
        return ""
    div_conteudo = span_rotulo.find_parent("div")
    if not div_conteudo:
        return ""
    texto = div_conteudo.find("div")
    return texto.get_text(strip=True) if texto else ""


def extrair_partes(
    html: str,
) -> Tuple[List[str], List[str], List[str], List[str], Dict[str, str]]:
    if not html:
        return [], [], [], [], {}

    soup = BeautifulSoup(html, "html.parser")
    tabela = soup.find("table", id="tablePartesPrincipais") or soup
    requerentes, devedores = [], []
    advogados_req, advogados_dev = [], []

    # estrutura para armazenar colunas dinâmicas: cada chave terá 'valores' e 'advogados'
    partes_em_colunas: Dict[str, Dict[str, List[str]]] = {}

    # Tipos que já são tratados separadamente (não criar colunas adicionais)
    tipos_ignorados = {
        "REQTE", "REQUERENTE", "EXEQUENTE", "PARTE ATIVA",
        "DEVEDOR", "DEVEDORA", "ENT. DEVEDORA", "REQUERIDO",
        "EXECUTADO", "PARTE PASSIVA"
    }

    for linha in tabela.find_all("tr"):
        celulas = linha.find_all(["td", "th"])
        if not celulas:
            continue

        # Ignora linhas de cabeçalho
        if all(celula.name == "th" for celula in celulas):
            continue

        tds = [celula for celula in celulas if celula.name == "td"]
        if not tds:
            continue

        # Extrai o tipo de participação
        span_tipo = linha.find("span", class_="tipoDeParticipacao")
        tipo_original = span_tipo.get_text(strip=True) if span_tipo else ""
        tipo = tipo_original.upper()

        # Extrai todas as informações da linha (texto de cada célula)
        informacoes_linha: List[str] = []
        for celula in tds:
            texto_celula = celula.get_text(separator=" ", strip=True)
            if texto_celula:
                informacoes_linha.append(texto_celula)

        # Processa nome e todos os advogados quando presentes
        td_nome = linha.find("td", class_="nomeParteEAdvogado")
        nome_parte = None
        advogados_local: List[str] = []
        if td_nome:
            # extrai o nome como o texto antes do primeiro <br> ou antes do primeiro span de rótulo
            nome_parts: List[str] = []
            for item in td_nome.contents:
                # para ao encontrar <br>
                if getattr(item, "name", None) == "br":
                    break
                # para ao encontrar rótulo (ex.: <span class="mensagemExibindo">Advogado:</span>)
                if getattr(item, "name", None) == "span" and "mensagemExibindo" in (item.get("class") or []):
                    break
                if isinstance(item, str):
                    t = item.strip()
                    if t:
                        nome_parts.append(t)
                else:
                    t = item.get_text(strip=True)
                    if t:
                        nome_parts.append(t)
            nome_parte = " ".join(nome_parts).strip() if nome_parts else None

            # coleta todos os advogados/advogadas listados na célula
            for span in td_nome.find_all("span", class_="mensagemExibindo"):
                texto_rotulo = (span.get_text() or "").strip()
                if "ADVOGAD" in texto_rotulo.upper():
                    # procura o próximo conteúdo relevante (pode ser texto direto ou uma tag)
                    for sib in span.next_siblings:
                        if isinstance(sib, str):
                            s = sib.strip()
                            if s:
                                advogados_local.append(s)
                                break
                        else:
                            s = sib.get_text(strip=True)
                            if s:
                                advogados_local.append(s)
                                break

        # Mantém lógica original para requerentes e devedores
        if nome_parte:
            if any(palavra in tipo for palavra in ["REQTE", "REQUERENTE", "EXEQUENTE", "PARTE ATIVA"]):
                requerentes.append(nome_parte)
                if advogados_local:
                    advogados_req.extend(advogados_local)
            elif any(palavra in tipo for palavra in [
                "DEVEDOR", "DEVEDORA", "ENT. DEVEDORA", "REQUERIDO", "EXECUTADO", "PARTE PASSIVA",
            ]):
                devedores.append(nome_parte)
                if advogados_local:
                    advogados_dev.extend(advogados_local)

        # Para tipos adicionais (não requerente/devedor), cria colunas dinâmicas
        if tipo_original and not any(palavra in tipo for palavra in tipos_ignorados):
            informacoes_limpas = [info for info in informacoes_linha if info.strip()]
            if informacoes_limpas:
                chave_coluna = informacoes_limpas[0]
                # Prefere o nome extraído (limpo) como valor; se não existir, usa o texto bruto das células
                if nome_parte:
                    valor_coluna = nome_parte
                else:
                    valor_coluna = " | ".join(informacoes_limpas[1:]) if len(informacoes_limpas) > 1 else ""

                if chave_coluna not in partes_em_colunas:
                    partes_em_colunas[chave_coluna] = {"valores": [], "advogados": []}

                # adiciona o valor (nome ou outras informações)
                if valor_coluna:
                    partes_em_colunas[chave_coluna]["valores"].append(valor_coluna)

                # adiciona advogados separados, se houver
                for a in advogados_local:
                    if a:
                        partes_em_colunas[chave_coluna]["advogados"].append(a)

    # formata o dicionário final: cria colunas para valores e para advogados separados
    partes_em_colunas_formatadas: Dict[str, str] = {}
    for chave, blocos in partes_em_colunas.items():
        valores = [v for v in blocos.get("valores", []) if v]
        advs = [a for a in blocos.get("advogados", []) if a]
        if valores:
            partes_em_colunas_formatadas[chave] = " | ".join(valores)
        if advs:
            partes_em_colunas_formatadas[f"{chave} - Advogados"] = " | ".join(advs)

    return requerentes, devedores, advogados_req, advogados_dev, partes_em_colunas_formatadas

def extrair_movimentacoes(html: str) -> str:
    if not html:
        return ""
    soup = BeautifulSoup(html, "html.parser")
    linhas = [linha.get_text(separator=" ", strip=True) for linha in soup.find_all("tr")]
    return "\n".join(filter(None, linhas))


def abrir_pasta_digital(driver: Chrome):
    clicar_com_retentativa(driver, (By.ID, "linkPasta"))
    time.sleep(1.5)
    driver.switch_to.window(driver.window_handles[-1])


def baixar_pdf(driver: Chrome) -> Optional[str]:
    clicar_com_retentativa(driver, (By.ID, "selecionarButton"), timeout=25)
    clicar_com_retentativa(driver, (By.ID, "salvarButton"), timeout=25)

    clicar_com_retentativa(driver, (By.ID, "opcao1"), timeout=15)

    botao_continuar = aguardar(driver, (By.ID, "botaoContinuar"), EC.visibility_of_element_located, 20)
    ActionChains(driver).move_to_element(botao_continuar).pause(0.5).click().perform()

    aguardar(driver, (By.ID, "msgAguarde"), EC.invisibility_of_element_located, 40)
    clicar_com_retentativa(driver, (By.ID, "btnDownloadDocumento"), timeout=25)

    return aguardar_download(PASTA_DOWNLOAD, TEMPO_DOWNLOAD)


def fechar_abas_extras(driver: Chrome):
    while len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[-1])
        driver.close()
        time.sleep(0.3)
    driver.switch_to.window(driver.window_handles[0])


def aguardar_download(pasta: Path, timeout: int = 60) -> Optional[str]:
    tempo_inicial = time.time()
    existentes = {arquivo.name for arquivo in pasta.glob("*.pdf")}

    while time.time() - tempo_inicial < timeout:
        for arquivo in pasta.glob("*.pdf"):
            if arquivo.name not in existentes:
                if not arquivo.name.endswith(".crdownload"):
                    print(f"PDF baixado: {arquivo.name}")
                    return arquivo.name
        time.sleep(1)

    print("Tempo limite atingido. Nenhum PDF encontrado.")
    return None


def construir_resultado(
    processo: str,
    classe: str,
    assunto: str,
    foro: str,
    vara: str,
    juiz: str,
    data_hora: str,
    controle: str,
    area: str,
    valor: str,
    outros_numeros: str,
    requerentes: List[str],
    advogados_req: List[str],
    devedores: List[str],
    advogados_dev: List[str],
    movimentacoes: str,
    peticoes: str,
    incidentes: str,
    apensos: str,
    audiencias: str,
    caminho_pdf: Optional[str],
    status: str = "OK",
    partes_em_colunas: Optional[Dict[str, str]] = None,
) -> Dict[str, str]:
    link_pdf = (
        f'=HYPERLINK("{Path(PASTA_DOWNLOAD, caminho_pdf)}", "{Path(PASTA_DOWNLOAD, caminho_pdf)}")'
        if caminho_pdf
        else "Não baixado"
    )

    resultado = {
        "numero_processo": processo,
        "Status": status,
        "Classe": classe,
        "Assunto": assunto,
        "Foro": foro,
        "Vara": vara,
        "Juiz": juiz,
        "Distribuicao": data_hora,
        "Controle": controle,
        "Area": area,
        "ValorAcao": valor,
        "Outros numeros": outros_numeros,
        "Requerente": ", ".join(requerentes),
        "ADVOGADOS REQUERENTE": ", ".join(advogados_req),
        "Devedor": ", ".join(devedores),
        "ADVOGADOS DEVEDOR": ", ".join(advogados_dev),
        "Movimentacoes": movimentacoes,
        "Petições diversas": peticoes,
        "Incidentes, acoes incidentais, recursos e execucoes de sentencas": incidentes,
        "Apensos, Entranhados e Unificados": apensos,
        "Audiencias": audiencias,
        "PDF": link_pdf,
    }

    if partes_em_colunas:
        resultado.update(partes_em_colunas)

    return resultado


def registrar_erro(processo: str, erro: Exception) -> Dict[str, str]:
    mensagem = f"{type(erro).__name__}: {erro}"
    try:
        logger.error("Falha no processo %s | %s", processo, mensagem)
    except Exception:
        pass
    print(f"Erro ao consultar {processo}: {mensagem}")
    return construir_resultado(
        processo=processo,
        classe="Erro ou não encontrado",
        assunto="Erro ou não encontrado",
        foro="Erro",
        vara="Erro",
        juiz="Erro",
        data_hora="Erro",
        controle="Erro",
        area="Erro",
        valor="Erro",
        outros_numeros="Erro",
        requerentes=[],
        advogados_req=[],
        devedores=[],
        advogados_dev=[],
        movimentacoes=mensagem,
        peticoes="Erro",
        incidentes="Erro",
        apensos="Erro",
        audiencias="Erro",
        caminho_pdf=None,
        status="ERRO",
        partes_em_colunas={},
    )


def processar_processo(driver: Chrome, processo: str) -> Dict[str, str]:
    parte1, parte3 = separar_numero_processo(processo)

    driver.get(URL_CONSULTA)
    preencher_campo(driver, (By.ID, "numeroDigitoAnoUnificado"), parte1)
    preencher_campo(driver, (By.ID, "foroNumeroUnificado"), parte3)

    clicar_com_retentativa(driver, (By.ID, "botaoConsultarProcessos"))
    aguardar_resultado_consulta(driver)

    html_pagina = driver.page_source
    outros_numeros = extrair_outros_numeros(html_pagina)

    classe = extrair_texto_por_id(driver, "classeProcesso")
    assunto = extrair_texto_por_id(driver, "assuntoProcesso")
    foro = extrair_texto_por_id(driver, "foroProcesso")
    vara = extrair_texto_por_id(driver, "varaProcesso")
    juiz = extrair_texto_por_id(driver, "juizProcesso")
    data_hora = extrair_texto_por_id(driver, "dataHoraDistribuicaoProcesso")
    controle = extrair_texto_por_id(driver, "numeroControleProcesso")
    area = extrair_texto_por_id(driver, "areaProcesso")
    valor = extrair_texto_por_id(driver, "valorAcaoProcesso")

    peticoes = extrair_texto_por_id(driver, "processoSemDiversas")
    incidentes = extrair_texto_por_id(driver, "processoSemIncidentes")
    apensos = extrair_texto_por_id(driver, "dadosApensosNaoDisponiveis")
    audiencias = extrair_texto_por_id(driver, "processoSemAudiencias")

    html_partes = extrair_html(driver, (By.ID, "tablePartesPrincipais"))
    (
        requerentes,
        devedores,
        advogados_req,
        advogados_dev,
        partes_em_colunas,
    ) = extrair_partes(html_partes)

    html_movimentacoes = extrair_html(driver, (By.ID, "tabelaUltimasMovimentacoes"))
    movimentacoes = extrair_movimentacoes(html_movimentacoes)

    caminho_pdf = None
    #try:
        #abrir_pasta_digital(driver)
        #caminho_pdf = baixar_pdf(driver)
    #finally:
        #fechar_abas_extras(driver)

    return construir_resultado(
        processo=processo,
        classe=classe,
        assunto=assunto,
        foro=foro,
        vara=vara,
        juiz=juiz,
        data_hora=data_hora,
        controle=controle,
        area=area,
        valor=valor,
        outros_numeros=outros_numeros,
        requerentes=requerentes,
        advogados_req=advogados_req,
        devedores=devedores,
        advogados_dev=advogados_dev,
        movimentacoes=movimentacoes,
        peticoes=peticoes,
        incidentes=incidentes,
        apensos=apensos,
        audiencias=audiencias,
        caminho_pdf=caminho_pdf,
        partes_em_colunas=partes_em_colunas,
    )


def obter_colunas_resultado() -> List[str]:
    """Retorna a lista de colunas na ordem correta para o Excel."""
    return [
        "numero_processo",
        "Status",
        "Classe",
        "Assunto",
        "Foro",
        "Vara",
        "Juiz",
        "Distribuicao",
        "Controle",
        "Area",
        "ValorAcao",
        "Outros numeros",
        "Requerente",
        "ADVOGADOS REQUERENTE",
        "Devedor",
        "ADVOGADOS DEVEDOR",
        "Movimentacoes",
        "Petições diversas",
        "Incidentes, acoes incidentais, recursos e execucoes de sentencas",
        "Apensos, Entranhados e Unificados",
        "Audiencias",
        "PDF",
    ]


def inicializar_arquivo_resultados(nome_arquivo: str):
    """Cria o arquivo Excel com os cabeçalhos se não existir."""
    if Path(nome_arquivo).exists():
        return
    
    # Cria um novo workbook usando openpyxl diretamente
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Adiciona os cabeçalhos
    colunas = obter_colunas_resultado()
    for col_idx, coluna in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=coluna)
        cell.font = Font(bold=True)
    
    # Salva o arquivo
    wb.save(nome_arquivo)
    wb.close()
    print(f"Arquivo de resultados criado: {nome_arquivo}")


def adicionar_resultado_ao_excel(resultado: Dict[str, str], nome_arquivo: str):
    """Adiciona uma linha ao arquivo Excel existente."""
    try:
        # Carrega o workbook existente (data_only=False para preservar fórmulas)
        wb = openpyxl.load_workbook(nome_arquivo, data_only=False)
        ws = wb.active
        
        # Garante que os cabeçalhos base existam
        cabecalhos_existentes = [
            celula.value for celula in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        colunas_base = obter_colunas_resultado()
        colunas = [valor for valor in cabecalhos_existentes if valor]

        if not colunas:
            colunas = colunas_base.copy()
            for idx, coluna in enumerate(colunas, start=1):
                ws.cell(row=1, column=idx, value=coluna)
        else:
            for coluna_base in colunas_base:
                if coluna_base not in colunas:
                    colunas.append(coluna_base)
                    cell = ws.cell(row=1, column=len(colunas), value=coluna_base)
                    cell.font = Font(bold=True)

        # Adiciona dinamicamente eventuais novas colunas provenientes das partes
        dynamic_keys = [k for k in resultado.keys() if k not in colunas]
        if dynamic_keys:
            try:
                # Se a coluna 'ADVOGADOS DEVEDOR' existir, insere as novas colunas logo depois dela
                if "ADVOGADOS DEVEDOR" in colunas:
                    insert_pos = colunas.index("ADVOGADOS DEVEDOR")  # 0-based index
                    # excel insertion position is 1-based; insert after the found column
                    excel_insert_at = insert_pos + 2
                    for chave in dynamic_keys:
                        # insere uma coluna vazia no arquivo para deslocar dados
                        ws.insert_cols(excel_insert_at)
                        cell = ws.cell(row=1, column=excel_insert_at, value=chave)
                        cell.font = Font(bold=True)
                        # atualiza lista de colunas em memória para refletir a nova ordem
                        colunas.insert(insert_pos + 1, chave)
                        insert_pos += 1
                        excel_insert_at += 1
                else:
                    # fallback: anexa no final
                    for chave in dynamic_keys:
                        colunas.append(chave)
                        cell = ws.cell(row=1, column=len(colunas), value=chave)
                        cell.font = Font(bold=True)
            except Exception:
                # se qualquer erro ao inserir colunas, anexa no final como fallback seguro
                for chave in dynamic_keys:
                    if chave not in colunas:
                        colunas.append(chave)
                        cell = ws.cell(row=1, column=len(colunas), value=chave)
                        cell.font = Font(bold=True)

        # Encontra a próxima linha vazia
        proxima_linha = ws.max_row + 1

        for col_idx, coluna in enumerate(colunas, start=1):
            valor = resultado.get(coluna, "")
            if valor is None:
                valor = ""
            ws.cell(row=proxima_linha, column=col_idx, value=valor)
        
        # Salva o arquivo
        wb.save(nome_arquivo)
        wb.close()
        
    except Exception as e:
        print(f"Erro ao salvar resultado no Excel: {e}")
        # Fallback: usa pandas como backup
        try:
            if Path(nome_arquivo).exists():
                df_existente = pd.read_excel(nome_arquivo)
                df_novo = pd.DataFrame([resultado])
                df_combinado = pd.concat([df_existente, df_novo], ignore_index=True)
            else:
                df_combinado = pd.DataFrame([resultado])
            df_combinado.to_excel(nome_arquivo, index=False, engine='openpyxl')
        except Exception as e2:
            print(f"Erro no fallback ao salvar: {e2}")


def main():
    # configura logger de erros
    logger.setLevel(logging.INFO)
    formato = logging.Formatter(
        fmt="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    try:
        handler = RotatingFileHandler(LOG_ARQUIVO, maxBytes=1_000_000, backupCount=3, encoding="utf-8")
    except Exception:
        handler = logging.FileHandler(LOG_ARQUIVO, encoding="utf-8")
    handler.setLevel(logging.INFO)
    handler.setFormatter(formato)
    if not logger.handlers:
        logger.addHandler(handler)

    processos = carregar_processos(CAMINHO_PLANILHA)
    if not processos:
        print("Nenhum número de processo encontrado na planilha.")
        return

    # Inicializa o arquivo de resultados com cabeçalhos
    inicializar_arquivo_resultados(NOME_ARQUIVO_RESULTADOS)
    
    driver: Optional[Chrome] = None

    try:
        driver = inicializar_driver()
        driver.get(URL_LOGIN)
        input("Faça o login manualmente e pressione ENTER para continuar...")

        for indice, processo in enumerate(processos, start=1):
            print(f"Processando {indice}/{len(processos)} - processo: {processo}")
            try:
                resultado = processar_processo(driver, processo)
            except (NoSuchElementException, TimeoutException, WebDriverException, ValueError) as erro:
                resultado = registrar_erro(processo, erro)
            
            # Salva o resultado imediatamente no Excel
            adicionar_resultado_ao_excel(resultado, NOME_ARQUIVO_RESULTADOS)
            print(f"Resultado salvo no Excel para o processo {processo}")
            # Pausa curta para estabilizar antes do próximo processo
            #time.sleep(GAP_ENTRE_PROCESSOS)
            # Pausa longa de 2 minutos conforme solicitado
            #time.sleep(PAUSA_2_MINUTOS)

    finally:
        if driver:
            driver.quit()
    
    print(f"\nProcessamento concluído! Todos os resultados foram salvos em {NOME_ARQUIVO_RESULTADOS}")


if __name__ == "__main__":
    main()