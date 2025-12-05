import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# Configurações
PASTA_PDFS = r"xxxxxxxx"
ARQUIVO_PROCESSOS = r"xxxxxxxx"
ARQUIVO_SAIDA = r"xxxxxxxx"


def buscar_caminhos_pdfs():
    """Busca os caminhos dos PDFs baseado nos números de processo"""
    
    # Ler o arquivo de processos
    try:
        df = pd.read_excel(ARQUIVO_PROCESSOS)
        print(f"[OK] Arquivo '{ARQUIVO_PROCESSOS}' carregado com sucesso!")
        print(f"[INFO] Total de processos: {len(df)}")
    except FileNotFoundError:
        print(f"[ERRO] Arquivo não encontrado: {ARQUIVO_PROCESSOS}")
        return
    except Exception as e:
        print(f"[ERRO] Erro ao ler o arquivo: {e}")
        return
    
    # Identificar a coluna com número do processo (primeira coluna ou coluna com "processo" no nome)
    coluna_processo = None
    for col in df.columns:
        if "processo" in str(col).lower() or "numero" in str(col).lower():
            coluna_processo = col
            break
    
    # Se não encontrou, usar a primeira coluna
    if coluna_processo is None:
        coluna_processo = df.columns[0]
    
    print(f"[INFO] Usando coluna: '{coluna_processo}'")
    
    # Criar listas para armazenar resultados
    resultados = []
    encontrados = 0
    nao_encontrados = 0
    
    # Verificar se a pasta de PDFs existe
    if not os.path.exists(PASTA_PDFS):
        print(f"[ERRO] Pasta de PDFs não encontrada: {PASTA_PDFS}")
        return
    
    # Listar todos os arquivos PDF na pasta (para busca mais eficiente)
    print(f"\n[...] Listando arquivos na pasta de PDFs...")
    try:
        arquivos_na_pasta = set(os.listdir(PASTA_PDFS))
        print(f"[OK] {len(arquivos_na_pasta)} arquivos encontrados na pasta")
    except Exception as e:
        print(f"[ERRO] Erro ao listar pasta: {e}")
        return
    
    print(f"\n[...] Buscando PDFs para cada processo...\n")
    
    # Buscar cada processo
    for idx, row in df.iterrows():
        numero_processo = str(row[coluna_processo]).strip()
        
        # Ignorar valores vazios ou inválidos
        if not numero_processo or numero_processo.lower() in ["nan", "none", ""]:
            resultados.append({
                "Número do Processo": numero_processo,
                "Caminho do PDF": "Número de processo inválido",
                "Status": "Inválido"
            })
            continue
        
        # Nome esperado do arquivo PDF
        nome_pdf = f"{numero_processo}.pdf"
        
        # Verificar se o arquivo existe
        if nome_pdf in arquivos_na_pasta:
            caminho_completo = os.path.join(PASTA_PDFS, nome_pdf)
            resultados.append({
                "Número do Processo": numero_processo,
                "Caminho do PDF": caminho_completo,
                "Status": "Encontrado"
            })
            encontrados += 1
            print(f"[OK] {numero_processo} -> Encontrado")
        else:
            resultados.append({
                "Número do Processo": numero_processo,
                "Caminho do PDF": "PDF não encontrado",
                "Status": "Não encontrado"
            })
            nao_encontrados += 1
            print(f"[X] {numero_processo} -> Não encontrado")
    
    # Criar DataFrame com resultados
    df_resultados = pd.DataFrame(resultados)
    
    # Salvar em Excel com formatação
    try:
        # Criar workbook com openpyxl para aplicar formatação
        wb = Workbook()
        ws = wb.active
        ws.title = "Caminhos PDFs"
        
        # Estilo para cabeçalhos em negrito
        fonte_negrito = Font(bold=True)
        
        # Escrever cabeçalhos
        colunas = ["Número do Processo", "Caminho do PDF", "Status"]
        for col_idx, coluna in enumerate(colunas, start=1):
            cell = ws.cell(row=1, column=col_idx, value=coluna)
            cell.font = fonte_negrito
        
        # Escrever dados
        for row_idx, resultado in enumerate(resultados, start=2):
            ws.cell(row=row_idx, column=1, value=resultado["Número do Processo"])
            ws.cell(row=row_idx, column=2, value=resultado["Caminho do PDF"])
            ws.cell(row=row_idx, column=3, value=resultado["Status"])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 15
        
        # Salvar
        wb.save(ARQUIVO_SAIDA)
        wb.close()
        
        print(f"\n{'='*60}")
        print(f"[CONCLUÍDO] Resultados salvos em: {ARQUIVO_SAIDA}")
        print(f"{'='*60}")
        print(f"  Total de processos: {len(resultados)}")
        print(f"  PDFs encontrados: {encontrados}")
        print(f"  PDFs não encontrados: {nao_encontrados}")
        print(f"{'='*60}")
        
    except PermissionError:
        print(f"\n[ERRO] Não foi possível salvar o arquivo!")
        print(f"Verifique se o arquivo '{ARQUIVO_SAIDA}' está aberto.")
    except Exception as e:
        print(f"\n[ERRO] Erro ao salvar: {e}")


if __name__ == "__main__":
    buscar_caminhos_pdfs()

